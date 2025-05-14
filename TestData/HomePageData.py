# from PythonSelFramework.TestData.excelDemo import dict1
import openpyxl


class HomePageData:
  test_homepage_params = [{'firstname': 'Mihai',
                           'lastname': 'Neagu',
                           'gender': 'Male'},
                          {'firstname': 'Anshika',
                           'lastname': 'Rahul',
                           'gender': 'Female'}]

  # test_excel_params = [dict1]

  @staticmethod  # avoid creation of objects
  def getTestData(test_case_name):  # integrate Excel utility into Selenium Python Framework

    excel_file = openpyxl.load_workbook('../PythonDemo.xlsx')
    sheet = excel_file.active
    dict3 = {}
    # test_case_name = sheet['a3'].value
    for i in range(1, sheet.max_row + 1):  # the range it starts from 1 serves excel annotation purpose
      if sheet.cell(row=i, column=1).value == test_case_name:
        for j in range(2, sheet.max_column + 1):  # build a Utility to retrieve values from Sheet based on conditions
          cell_value = sheet.cell(row=i, column=j).value
          dict3[sheet.cell(row=1, column=j).value] = cell_value
    return [dict3]  # this ensures pytest correctly passes the dictionary as a parameter
    # instead of trying to iterate over dictionary keys
  # makes it a list, ensuring pytest treats each dictionary inside it as a test case.

  # pytest.fixture(params=HomePageData.getTestData('Testcase1'))
  # expects multiple test cases as an iterable (list or tuple)
