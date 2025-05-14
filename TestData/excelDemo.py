import openpyxl
from openpyxl.utils import get_column_letter as let

file = openpyxl.load_workbook('../PythonDemo.xlsx')
sheet = file.active
dict1 = {}
dict2 = {}
cell1 = sheet.cell(row=1, column=2)  # target the specific cell
print(cell1.value)  # extract value from the cell
sheet.cell(row=2, column=2).value = 'SuperMike'  # assign value to specific cell
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)  # check how many rows are in the active sheet
print(sheet.max_column)  # numbers of the columns
print(sheet['a7'].value)  # print the value for cell 'A7'

for i in range(1, sheet.max_row + 1):  # the range it starts from 1 serves excel annotation purpose
  if sheet.cell(row=i, column=1).value == sheet['a3'].value:
    for j in range(2, sheet.max_column + 1):  # build a Utility to retrieve values from Sheet based on conditions
      cell_value = sheet.cell(row=i, column=j).value
      dict1[sheet.cell(row=1, column=j).value] = cell_value
      # print(f'Row {i}, Column {j}: {cell_value}')
print(dict1)

# for k in range(1, sheet.max_row + 1):  # update the row
#   for l in range(1, sheet.max_column + 1):  # update the column
#     m = let(l)
#     cell_value2 = sheet[f'{m}{k}'].value
#     dict2[sheet.cell(row=1, column=l).value] = cell_value2
#     print(f'Row {k}, Column{m}: {cell_value2}')
#     print(dict2)
